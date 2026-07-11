#!/usr/bin/env python3
"""Build a plant-only PMG service snapshot from Katie's workbook feed.

The importer is intentionally conservative:
- only Plant tabs plus Tractor/Kubota Mower/Forklift are read;
- non-plant tabs are ignored and their sheet names are not written into app data;
- rows are matched to plant numbers where possible;
- unclear schedule/service gaps are put into review for Bill/Katie/Tony.
"""

from __future__ import annotations

import argparse
import calendar
import datetime as dt
import hashlib
import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_SERVICE_WORKBOOK = Path("/Users/bill/Desktop/HGV, VEHICLE & PLANT SERVICE SHEET.xlsx")
REVISED_SCHEDULE_WORKBOOK = Path("/Users/bill/.openclaw/workspace/shared-docs/PMG/Plant Schedule revised 2024.xlsx")
LEGACY_SCHEDULE_WORKBOOK = Path(
    "/Users/bill/Library/CloudStorage/OneDrive-PMGroundworksLimited/SHARED/Plant Schedule.xlsx"
)
PLANT_SCHEDULE_SHAREPOINT_URL = (
    "https://pmgroundworksblackpool.sharepoint.com/:x:/r/sites/PMGroundWorksTeamSite/"
    "_layouts/15/Doc.aspx?sourcedoc=%7B3C4E7077-B31D-471E-AE5A-C026175BFF75%7D"
    "&file=Plant%20Schedule%20revised%202024.xlsx&action=default&mobileredirect=true&DefaultItemOpen=1"
)
DEFAULT_SHAREPOINT_ROOT = "SHARED/Plant and Tool Service Records"
DEFAULT_WORKER_URL = "https://pmg-driver-sync.jimpmgr.workers.dev"
USER_AGENT = "Mozilla/5.0 PMG-Plant-Workbook-Import/1.0"

PLANT_SHEET_RE = re.compile(r"^plant\s*(\d+)$", re.I)
SPECIAL_PLANT_SHEETS = {
    "tractor": ("plant-tractor", "Tractor", "TRACTOR"),
    "kubota mower": ("plant-kubota-mower", "Kubota Mower", "KUBOTA MOWER"),
    "forklift": ("plant-forklift", "Forklift", "FORKLIFT"),
}
FILTER_TERMS = ("filter", "belt", "activ 8", "activ8", "hydraulic", "engine oil", "fuel", "oil")
SERVICE_TERMS = ("service", "serviced", "oil", "filter", "belt", "activ8", "activ 8")
REPAIR_TERMS = ("hose", "pump", "screen", "battery", "gasket", "burst", "repair", "replace", "fitted")
UNUSED_PLANT_TERMS = ("spare plant", "unused", "not used")
SERVICE_INTERVAL_MONTHS = 12


def default_schedule_workbook() -> Path:
    return REVISED_SCHEDULE_WORKBOOK if REVISED_SCHEDULE_WORKBOOK.exists() else LEGACY_SCHEDULE_WORKBOOK


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\u2019", "'")).strip()


def parse_date(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = clean_cell(value)
    if not text:
        return None
    iso_match = re.search(r"\b(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})\b", text)
    if iso_match:
        year, month, day = map(int, iso_match.groups())
        return safe_date(year, month, day)
    text = text.replace("/.", ".").replace("./", ".").replace("//", "/")
    uk_match = re.search(r"\b(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\b", text)
    if uk_match:
        day, month, year = map(int, uk_match.groups())
        if year < 100:
            year += 2000 if year < 70 else 1900
        return safe_date(year, month, day)
    return None


def safe_date(year: int, month: int, day: int) -> str | None:
    try:
        return dt.date(year, month, day).isoformat()
    except ValueError:
        return None


def add_months(value: str | None, months: int) -> str | None:
    if not value:
        return None
    try:
        date = dt.date.fromisoformat(value)
    except ValueError:
        return None
    month_index = date.month - 1 + months
    year = date.year + month_index // 12
    month = month_index % 12 + 1
    day = min(date.day, calendar.monthrange(year, month)[1])
    return dt.date(year, month, day).isoformat()


def parse_amount(value: Any) -> float | None:
    text = clean_cell(value).replace(",", "")
    if not text:
        return None
    match = re.fullmatch(r"-?\d+(\.\d{1,2})?", text)
    if not match:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_hours(text: str) -> int | None:
    match = re.search(r"\b(\d{2,6})(?:\.\d+)?\s*(?:hours?|hrs?|hr)\b", text, re.I)
    if match:
        return int(match.group(1))
    return None


def stable_id(*parts: Any) -> str:
    joined = "|".join(clean_cell(part).lower() for part in parts)
    return hashlib.sha1(joined.encode("utf-8")).hexdigest()[:16]


def asset_id_from_sheet(sheet_name: str) -> tuple[str, str, str] | None:
    match = PLANT_SHEET_RE.match(sheet_name.strip())
    if match:
        number = str(int(match.group(1)))
        return f"plant-{number}", number, f"Plant {number}"
    return SPECIAL_PLANT_SHEETS.get(sheet_name.strip().lower())


def evidence_path_for(asset: dict[str, Any], sharepoint_root: str) -> str:
    label = asset.get("displayName") or asset.get("plantNumber") or asset["id"]
    safe_label = re.sub(r"[\\/]+", "-", str(label)).strip()
    return f"{sharepoint_root}/{safe_label}"


def is_unused_schedule_row(cells: list[str]) -> bool:
    row_text = " ".join(cells).lower()
    has_no_identity = not cells[2] and not cells[3]
    return has_no_identity and any(term in row_text for term in UNUSED_PLANT_TERMS)


def schedule_assets(path: Path, sharepoint_root: str) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[str]]:
    assets: dict[str, dict[str, Any]] = {}
    review: list[dict[str, Any]] = []
    skipped_unused: list[str] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    serials: dict[str, list[str]] = {}

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = [clean_cell(v) for v in row[:10]]
        if not any(cells):
            continue
        number = cells[0]
        if not re.fullmatch(r"\d+", number):
            review.append(
                {
                    "id": f"review-{stable_id('schedule', row_index, cells)}",
                    "type": "schedule_row_unmatched",
                    "severity": "review",
                    "message": "Plant schedule row does not have a numeric plant number.",
                    "source": {"workbook": str(path), "row": row_index},
                    "raw": cells,
                }
            )
            continue

        plant_number = str(int(number))
        if is_unused_schedule_row(cells):
            skipped_unused.append(plant_number)
            continue

        asset_id = f"plant-{plant_number}"
        notes = [cell for cell in cells[5:] if cell]
        asset = {
            "id": asset_id,
            "plantNumber": plant_number,
            "displayName": f"Plant {plant_number} - {cells[2]}" if cells[2] else f"Plant {plant_number}",
            "machineType": cells[2],
            "serialNumber": cells[3],
            "year": "",
            "site": cells[4],
            "scheduleInspectionExpiry": parse_date(row[1]) if len(row) > 1 else None,
            "scheduleNotes": notes,
            "sharePointEvidencePath": "",
            "sources": [{"type": "plant_schedule", "workbook": str(path), "row": row_index}],
        }
        asset["sharePointEvidencePath"] = evidence_path_for(asset, sharepoint_root)
        assets[asset_id] = asset
        if cells[3]:
            serials.setdefault(cells[3].upper(), []).append(asset_id)
        if notes:
            review.append(
                {
                    "id": f"review-{stable_id('schedule-note', asset_id, row_index, notes)}",
                    "type": "schedule_note",
                    "severity": "note",
                    "assetId": asset_id,
                    "plantNumber": plant_number,
                    "message": "; ".join(notes),
                    "source": {"workbook": str(path), "row": row_index},
                }
            )

    for serial, ids in serials.items():
        if len(ids) > 1:
            review.append(
                {
                    "id": f"review-{stable_id('duplicate-serial', serial, ids)}",
                    "type": "duplicate_serial",
                    "severity": "review",
                    "assetIds": ids,
                    "message": f"Serial number {serial} appears on more than one plant record.",
                }
            )

    return assets, review, skipped_unused


def merged_schedule_assets(
    schedule_path: Path,
    sharepoint_root: str,
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[str], list[str]]:
    assets, review, skipped_unused = schedule_assets(schedule_path, sharepoint_root)
    sources = [str(schedule_path)]
    include_legacy = os.environ.get("PMG_PLANT_INCLUDE_LEGACY_SCHEDULE", "").strip() == "1"
    if include_legacy and schedule_path.resolve() == REVISED_SCHEDULE_WORKBOOK.resolve() and LEGACY_SCHEDULE_WORKBOOK.exists():
        try:
            legacy_assets, _legacy_review, legacy_skipped_unused = schedule_assets(LEGACY_SCHEDULE_WORKBOOK, sharepoint_root)
        except (OSError, PermissionError) as exc:
            review.append(
                {
                    "id": f"review-{stable_id('legacy-schedule-unavailable', LEGACY_SCHEDULE_WORKBOOK, type(exc).__name__)}",
                    "type": "legacy_schedule_unavailable",
                    "severity": "review",
                    "message": (
                        "Legacy Plant Schedule.xlsx could not be read; continuing with "
                        "Plant Schedule revised 2024.xlsx only."
                    ),
                    "source": {"workbook": str(LEGACY_SCHEDULE_WORKBOOK)},
                    "error": str(exc)[:300],
                }
            )
            legacy_assets = {}
            legacy_skipped_unused = []
        skipped_unused.extend(number for number in legacy_skipped_unused if number not in skipped_unused)
        current_unused_numbers = set(skipped_unused)
        if legacy_assets or legacy_skipped_unused:
            sources.append(str(LEGACY_SCHEDULE_WORKBOOK))
        for asset_id, asset in legacy_assets.items():
            if str(asset.get("plantNumber", "")) in current_unused_numbers:
                continue
            if asset_id in assets:
                continue
            fallback = dict(asset)
            fallback["site"] = ""
            fallback["sources"] = [
                *fallback.get("sources", []),
                {"type": "legacy_schedule_fallback_missing_from_revised", "workbook": str(LEGACY_SCHEDULE_WORKBOOK)},
            ]
            fallback["sharePointEvidencePath"] = evidence_path_for(fallback, sharepoint_root)
            assets[asset_id] = fallback
            review.append(
                {
                    "id": f"review-{stable_id('legacy-schedule-fallback', asset_id)}",
                    "type": "missing_from_revised_schedule",
                    "severity": "review",
                    "assetId": asset_id,
                    "plantNumber": fallback.get("plantNumber", ""),
                    "message": (
                        f"Plant {fallback.get('plantNumber', '')} is not in Plant Schedule revised 2024.xlsx; "
                        "kept from the legacy schedule with no current location."
                    ),
                    "source": {"workbook": str(LEGACY_SCHEDULE_WORKBOOK)},
                }
            )
    return assets, review, sources, skipped_unused


def row_texts(row: tuple[Any, ...]) -> list[str]:
    return [clean_cell(value) for value in row if clean_cell(value)]


def is_context_row(texts: list[str], joined_lower: str) -> bool:
    if len(texts) != 1:
        return False
    text = texts[0]
    if parse_date(text) or parse_hours(text):
        return False
    if any(term in joined_lower for term in FILTER_TERMS):
        return False
    if "inspection" in joined_lower:
        return False
    if text.lower().startswith("plant "):
        return False
    return True


def event_category(joined_lower: str, context: str, amount: float | None) -> str:
    context_lower = context.lower()
    if "inspection" in joined_lower:
        return "inspection"
    if any(term in joined_lower or term in context_lower for term in SERVICE_TERMS):
        return "service"
    if any(term in joined_lower for term in REPAIR_TERMS):
        return "repair"
    if amount is not None:
        return "invoice"
    return "note"


def looks_like_money_token(value: str) -> bool:
    text = clean_cell(value).replace(",", "")
    return "£" in text or bool(re.fullmatch(r"-?\d+\.\d{1,2}", text))


def looks_like_workbook_reference(value: str) -> bool:
    text = clean_cell(value)
    return bool(re.fullmatch(r"[A-Za-z]{1,6}-?\d{4,}[A-Za-z0-9-]*", text) or re.fullmatch(r"\d{4,}", text))


def public_filter_note(texts: list[str]) -> str:
    visible: list[str] = []
    for text in texts:
        if parse_date(text) or parse_hours(text) or looks_like_money_token(text) or looks_like_workbook_reference(text):
            continue
        visible.append(text)
    if visible:
        return " | ".join(visible)
    return " | ".join(text for text in texts if not looks_like_money_token(text))


def extract_filter_parts(texts: list[str], asset_id: str, source: dict[str, Any]) -> list[dict[str, Any]]:
    parts: list[dict[str, Any]] = []
    if not texts:
        return parts
    joined = " | ".join(texts)
    if not any(term in joined.lower() for term in FILTER_TERMS):
        return parts

    # Common format: "Oil Filter", "W21ES01600".
    if len(texts) >= 2 and any(term in texts[0].lower() for term in FILTER_TERMS):
        label = texts[0].rstrip(":")
        for code in texts[1:]:
            if parse_hours(code):
                continue
            if parse_date(code) or looks_like_money_token(code):
                continue
            if any(term in code.lower() for term in ("hours", "service:")):
                continue
            parts.append(
                {
                    "id": f"part-{stable_id(asset_id, source.get('row'), label, code)}",
                    "assetId": asset_id,
                    "type": label,
                    "code": code,
                    "notes": public_filter_note(texts),
                    "source": source,
                }
            )
        if parts:
            return parts

    # Fallback: preserve the whole row as a parts/service clue.
    parts.append(
        {
            "id": f"part-{stable_id(asset_id, source.get('row'), joined)}",
            "assetId": asset_id,
            "type": "Service/parts note",
            "code": "",
            "notes": public_filter_note(texts),
            "source": source,
        }
    )
    return parts


def ensure_service_asset(
    assets: dict[str, dict[str, Any]],
    asset_id: str,
    plant_number: str,
    display_name: str,
    first_title: str,
    sharepoint_root: str,
    source: dict[str, Any],
) -> None:
    if asset_id in assets:
        assets[asset_id].setdefault("sources", []).append(source)
        return
    machine_type = first_title if first_title and not first_title.lower().startswith("plant ") else ""
    asset = {
        "id": asset_id,
        "plantNumber": plant_number,
        "displayName": display_name if not machine_type else f"{display_name} - {machine_type}",
        "machineType": machine_type,
        "serialNumber": "",
        "year": "",
        "site": "",
        "scheduleInspectionExpiry": None,
        "scheduleNotes": [],
        "sharePointEvidencePath": "",
        "sources": [source],
    }
    asset["sharePointEvidencePath"] = evidence_path_for(asset, sharepoint_root)
    assets[asset_id] = asset


def service_feed(
    path: Path,
    assets: dict[str, dict[str, Any]],
    sharepoint_root: str,
) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]], list[str], set[str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    history: list[dict[str, Any]] = []
    filters: dict[str, list[dict[str, Any]]] = {}
    review: list[dict[str, Any]] = []
    ignored_sheets: list[str] = []
    matched_asset_ids: set[str] = set()

    for sheet_name in wb.sheetnames:
        sheet_match = asset_id_from_sheet(sheet_name)
        if not sheet_match:
            ignored_sheets.append(sheet_name)
            continue
        asset_id, plant_number, display_name = sheet_match
        ws = wb[sheet_name]
        title_cells = row_texts(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        first_title = title_cells[0] if title_cells else display_name
        source_ref = {"type": "service_workbook", "workbook": str(path), "sheet": sheet_name}
        existed = asset_id in assets
        ensure_service_asset(assets, asset_id, plant_number, display_name, first_title, sharepoint_root, source_ref)
        matched_asset_ids.add(asset_id)

        if not existed:
            review.append(
                {
                    "id": f"review-{stable_id('service-tab-no-schedule', asset_id, sheet_name)}",
                    "type": "service_tab_no_schedule_asset",
                    "severity": "review",
                    "assetId": asset_id,
                    "plantNumber": plant_number,
                    "message": f"{sheet_name} exists in the service workbook but is not on the plant schedule.",
                    "source": source_ref,
                }
            )

        context = ""
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            texts = row_texts(row)
            if not texts:
                continue
            joined = " | ".join(texts)
            joined_lower = joined.lower()
            source = {**source_ref, "row": row_index}

            if row_index == 1:
                continue
            if "inspection" in joined_lower:
                inspection_date = parse_date(joined)
                if inspection_date:
                    assets[asset_id]["lastInspectionDate"] = max(
                        filter(None, [assets[asset_id].get("lastInspectionDate"), inspection_date])
                    )
                    history.append(
                        {
                            "id": f"hist-{stable_id(asset_id, 'inspection', row_index, joined)}",
                            "assetId": asset_id,
                            "plantNumber": plant_number,
                            "date": inspection_date,
                            "category": "inspection",
                            "sourceName": "Katie workbook",
                            "description": joined,
                            "hours": parse_hours(joined),
                            "invoiceNumber": "",
                            "amount": None,
                            "source": source,
                        }
                    )
                else:
                    review.append(
                        {
                            "id": f"review-{stable_id(asset_id, 'inspection-no-date', row_index, joined)}",
                            "type": "inspection_date_unclear",
                            "severity": "review",
                            "assetId": asset_id,
                            "plantNumber": plant_number,
                            "message": joined,
                            "source": source,
                        }
                    )
                continue

            if is_context_row(texts, joined_lower):
                context = texts[0].rstrip(":")
                continue

            filters.setdefault(asset_id, [])
            for part in extract_filter_parts(texts, asset_id, source):
                if all(existing["id"] != part["id"] for existing in filters[asset_id]):
                    filters[asset_id].append(part)

            date = parse_date(row[0]) or parse_date(joined)
            hours = parse_hours(joined)
            if hours:
                assets[asset_id]["lastKnownHours"] = max(assets[asset_id].get("lastKnownHours") or 0, hours)

            if not date:
                if hours or any(term in joined_lower for term in SERVICE_TERMS):
                    review.append(
                        {
                            "id": f"review-{stable_id(asset_id, 'service-row-no-date', row_index, joined)}",
                            "type": "service_row_needs_date",
                            "severity": "review",
                            "assetId": asset_id,
                            "plantNumber": plant_number,
                            "message": joined,
                            "source": source,
                        }
                    )
                continue

            invoice_number = clean_cell(row[1]) if len(row) > 1 else ""
            amount_cell_index = 2
            amount = parse_amount(row[amount_cell_index]) if len(row) > amount_cell_index else None
            if (
                not invoice_number
                and len(row) > 3
                and clean_cell(row[2])
                and parse_amount(row[3]) is not None
            ):
                invoice_number = clean_cell(row[2])
                amount_cell_index = 3
                amount = parse_amount(row[3])
            description_cells = [clean_cell(v) for v in row[amount_cell_index + 1 :] if clean_cell(v)]
            if not description_cells:
                amount_cell = clean_cell(row[amount_cell_index]) if len(row) > amount_cell_index else ""
                description_cells = [text for text in texts[1:] if text not in {invoice_number, amount_cell}]
            description = " | ".join(description_cells).strip(" |") or joined
            category = event_category(joined_lower, context, amount)
            event = {
                "id": f"hist-{stable_id(asset_id, row_index, date, joined)}",
                "assetId": asset_id,
                "plantNumber": plant_number,
                "date": date,
                "category": category,
                "sourceName": context or "Katie workbook",
                "description": description,
                "hours": hours,
                "invoiceNumber": invoice_number,
                "amount": amount,
                "source": source,
            }
            history.append(event)
            if category == "service":
                assets[asset_id]["lastServiceDate"] = max(filter(None, [assets[asset_id].get("lastServiceDate"), date]))
                if hours:
                    assets[asset_id]["lastServiceHours"] = hours
            if category in {"invoice", "repair", "service"}:
                assets[asset_id]["lastCostDate"] = max(filter(None, [assets[asset_id].get("lastCostDate"), date]))
            if description == joined and len(texts) <= 2:
                review.append(
                    {
                        "id": f"review-{stable_id(asset_id, 'dated-row-unclear', row_index, joined)}",
                        "type": "dated_row_needs_description",
                        "severity": "review",
                        "assetId": asset_id,
                        "plantNumber": plant_number,
                        "message": joined,
                        "source": source,
                    }
                )

    return history, filters, review, ignored_sheets, matched_asset_ids


def build_snapshot(schedule_path: Path, service_path: Path, sharepoint_root: str) -> dict[str, Any]:
    assets, schedule_review, schedule_sources, skipped_unused_numbers = merged_schedule_assets(schedule_path, sharepoint_root)
    history, filters, service_review, ignored_sheets, matched_asset_ids = service_feed(service_path, assets, sharepoint_root)

    for asset_id, asset in assets.items():
        if asset_id.startswith("plant-") and asset["plantNumber"].isdigit() and asset_id not in matched_asset_ids:
            service_review.append(
                {
                    "id": f"review-{stable_id('schedule-no-service-tab', asset_id)}",
                    "type": "schedule_asset_no_service_tab",
                    "severity": "note",
                    "assetId": asset_id,
                    "plantNumber": asset["plantNumber"],
                    "message": f"Plant {asset['plantNumber']} is on the plant schedule but has no plant tab in Katie's service workbook.",
                    "source": {"workbook": str(schedule_path)},
                }
            )

    service_date_schedule_count = 0
    for asset in assets.values():
        if asset.get("lastServiceDate") and not asset.get("nextServiceDueDate"):
            next_due = add_months(asset.get("lastServiceDate"), SERVICE_INTERVAL_MONTHS)
            if next_due:
                asset["nextServiceDueDate"] = next_due
                asset["nextServiceDueBasis"] = "12_month_service_rule"
                service_date_schedule_count += 1

    assets_list = sorted(
        assets.values(),
        key=lambda item: (0, int(item["plantNumber"])) if str(item.get("plantNumber", "")).isdigit() else (1, str(item.get("plantNumber", ""))),
    )
    history.sort(key=lambda item: (item.get("date") or "", item.get("plantNumber") or "", item.get("id") or ""), reverse=True)
    review_items = schedule_review + service_review
    review_items.sort(key=lambda item: (item.get("severity") != "review", item.get("type", ""), item.get("plantNumber", "")))

    return {
        "schema": "pmg-plant-snapshot-v1",
        "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "sharePointEvidenceRoot": sharepoint_root,
        "sources": {
            "plantScheduleWorkbook": str(schedule_path),
            "plantScheduleWorkbooks": schedule_sources,
            "plantScheduleSharePointUrl": PLANT_SCHEDULE_SHAREPOINT_URL,
            "serviceWorkbook": str(service_path),
            "serviceWorkbookMode": "local_sharepoint_sync_ready" if "Library/CloudStorage" in str(service_path) else "local_file_feed",
        },
        "plantOnlyRules": {
            "includedSheetPattern": "^Plant <number>$ plus Tractor, Kubota Mower and Forklift",
            "ignoredSheetCount": len(ignored_sheets),
            "ignoredSheetsStored": False,
            "unusedSchedulePlantNumbersSkipped": skipped_unused_numbers,
            "unusedSchedulePlantNumbersStoredAsAssets": False,
        },
        "assets": assets_list,
        "history": history,
        "filterParts": filters,
        "reviewItems": review_items,
        "stats": {
            "assetCount": len(assets_list),
            "historyEventCount": len(history),
            "filterPartCount": sum(len(parts) for parts in filters.values()),
            "reviewItemCount": len(review_items),
            "ignoredSheetCount": len(ignored_sheets),
            "serviceDateScheduleCount": service_date_schedule_count,
            "unusedSchedulePlantNumberCount": len(skipped_unused_numbers),
        },
    }


def worker_headers(api_key: str) -> dict[str, str]:
    return {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": USER_AGENT,
        "X-PMG-Key": api_key,
    }


def check_worker_routes(worker_url: str, api_key: str) -> dict[str, Any]:
    url = worker_url.rstrip("/") + "/plant/assets"
    request = urllib.request.Request(url, headers=worker_headers(api_key), method="GET")
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            payload = response.read().decode("utf-8", errors="replace")
            data = json.loads(payload) if payload else {}
            return {
                "ok": response.status == 200,
                "status": "ready" if response.status == 200 else "unexpected_status",
                "httpStatus": response.status,
                "assetCount": len(data.get("assets") or []) if isinstance(data, dict) else None,
                "schema": data.get("schema") if isinstance(data, dict) else "",
            }
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace").strip()
        if exc.code == 404:
            status = "not_deployed"
        elif exc.code == 401:
            status = "unauthorized"
        elif exc.code == 403:
            status = "forbidden"
        else:
            status = "http_error"
        return {"ok": False, "status": status, "httpStatus": exc.code, "detail": detail[:500]}
    except Exception as exc:
        return {"ok": False, "status": "request_failed", "detail": str(exc)[:500]}


def push_snapshot(snapshot: dict[str, Any], worker_url: str, api_key: str) -> dict[str, Any]:
    url = worker_url.rstrip("/") + "/plant/import-snapshot"
    body = json.dumps(snapshot, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        method="PUT",
        headers=worker_headers(api_key),
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = response.read().decode("utf-8")
        return json.loads(payload) if payload else {"ok": True}


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Import PMG plant service data from Katie's workbook feed.")
    parser.add_argument("--service-workbook", type=Path, default=DEFAULT_SERVICE_WORKBOOK)
    parser.add_argument("--schedule-workbook", type=Path, default=default_schedule_workbook())
    parser.add_argument("--sharepoint-root", default=DEFAULT_SHAREPOINT_ROOT)
    parser.add_argument("--out", type=Path, help="Write the snapshot JSON to this path.")
    parser.add_argument("--push", action="store_true", help="Push snapshot to the PMG worker /plant/import-snapshot endpoint.")
    parser.add_argument("--check-worker", action="store_true", help="Check whether the live PMG worker has the plant routes deployed.")
    parser.add_argument("--worker-url", default=DEFAULT_WORKER_URL)
    parser.add_argument("--api-key", default="")
    parser.add_argument("--dry-run", action="store_true", help="Print only the summary stats and review counts.")
    args = parser.parse_args(argv)

    if args.check_worker:
        if not args.api_key:
            parser.error("--api-key is required with --check-worker")
        status = check_worker_routes(args.worker_url, args.api_key)
        print(json.dumps(status, ensure_ascii=False, indent=2))
        return 0 if status.get("ok") else 2

    snapshot = build_snapshot(args.schedule_workbook, args.service_workbook, args.sharepoint_root)

    if args.out:
        args.out.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if args.push:
        if not args.api_key:
            parser.error("--api-key is required with --push")
        readiness = check_worker_routes(args.worker_url, args.api_key)
        if not readiness.get("ok"):
            print(json.dumps({"ok": False, "error": "worker_routes_not_ready", "routeStatus": readiness}, ensure_ascii=False, indent=2))
            return 2
        try:
            result = push_snapshot(snapshot, args.worker_url, args.api_key)
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace").strip()
            print(json.dumps({"ok": False, "error": f"push_failed_http_{exc.code}", "detail": detail[:500]}, ensure_ascii=False, indent=2))
            return 1
        except Exception as exc:
            print(json.dumps({"ok": False, "error": "push_failed", "detail": str(exc)}, ensure_ascii=False, indent=2))
            return 1
        print(json.dumps(result, ensure_ascii=False, indent=2))
    elif args.dry_run or not args.out:
        print(json.dumps({"stats": snapshot["stats"], "plantOnlyRules": snapshot["plantOnlyRules"]}, ensure_ascii=False, indent=2))

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
